import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Zero Traffic Monitor", layout="centered")

st.title("🟩 Zero Traffic Monitor")
st.markdown("*Upload your files below. The tool will filter cells with at least one day of zero traffic and return an Excel file.*")

# File Uploads
on_air_tracker = st.file_uploader("Upload On-Air Tracker", type=["xlsx", "xls", "csv"])
kpi_day1 = st.file_uploader("Upload KPI Day 1", type=["xlsx", "xls", "csv"])
kpi_day2 = st.file_uploader("Upload KPI Day 2", type=["xlsx", "xls", "csv"])
kpi_day3 = st.file_uploader("Upload KPI Day 3", type=["xlsx", "xls", "csv"])

def read_file(file):
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    else:
        return pd.read_excel(file)

if st.button("📊 Process Data"):
    if not all([on_air_tracker, kpi_day1, kpi_day2, kpi_day3]):
        st.error("Please upload all 4 required files.")
    else:
        try:
            tracker = read_file(on_air_tracker)
            kpi_files = [read_file(f) for f in [kpi_day1, kpi_day2, kpi_day3]]

            tracker.columns = tracker.columns.str.strip()
            tracker = tracker.rename(columns={"Logical Site ID": "Site Id"})

            if "Site Id" not in tracker.columns or "Site IP" not in tracker.columns:
                st.error("On-air tracker must contain 'Logical Site ID' and 'Site IP' columns.")
            else:
                output_df = pd.DataFrame()

                for i, df in enumerate(kpi_files):
                    df.columns = df.columns.str.strip()
                    date = df["Date"].iloc[0] if "Date" in df.columns else f"Day{i+1}"

                    if not all(col in df.columns for col in ["4G Cell Name", "Site Id", "Data Volume - Total (GB)"]):
                        st.error(f"KPI Day {i+1} must have '4G Cell Name', 'Site Id', and 'Data Volume - Total (GB)' columns.")
                        st.stop()

                    df["Site Id"] = df["Site Id"].astype(str).str.strip()
                    df["4G Cell Name"] = df["4G Cell Name"].astype(str).str.strip()

                    # Use 4G Cell Name as unique key
                    sub_df = df[["4G Cell Name", "Site Id", "Data Volume - Total (GB)"]].copy()
                    sub_df = sub_df.rename(columns={
                        "4G Cell Name": "4G Cell",
                        "Data Volume - Total (GB)": str(date)
                    })

                    if output_df.empty:
                        output_df = sub_df
                    else:
                        output_df = pd.merge(output_df, sub_df, on=["4G Cell", "Site Id"], how="outer")

                # Add IP_ID using Site Id
                tracker["Site Id"] = tracker["Site Id"].astype(str).str.strip()
                output_df["IP_ID"] = output_df["Site Id"].map(tracker.set_index("Site Id")["Site IP"].to_dict())

                # Rename Site Id to match the old column name
                output_df = output_df.rename(columns={"Site Id": "TCS_Logical_ID"})

                # Keep only rows with at least one day of zero or less traffic
                date_columns = output_df.columns.difference(["TCS_Logical_ID", "IP_ID", "4G Cell"])
                output_df = output_df[output_df[date_columns].le(0).any(axis=1)]

                # Remove rows with blank IP_ID
                output_df = output_df[output_df["IP_ID"].notna() & (output_df["IP_ID"].astype(str).str.strip() != "")]

                # Count unique IPs
                unique_ip_count = output_df["IP_ID"].nunique()

                # Save to Excel in memory
                output_buffer = BytesIO()
                with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                    output_df.to_excel(writer, index=False)
                    workbook = writer.book
                    worksheet = writer.sheets["Sheet1"]

                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    for cell in worksheet[1]:
                        cell.fill = yellow_fill

                st.success(f"✅ Total sites having Zero traffic is {unique_ip_count}")
                st.download_button(
                    "📥 Download Output Excel",
                    data=output_buffer.getvalue(),
                    file_name="zero_traffic_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"Error during processing: {e}")
